import os
import pickle
from datetime import datetime, timedelta
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import re
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

# --- 엑셀 데이터 읽기 ---
def open_edsd_score_excel(excel_path):
    if os.path.exists(excel_path):
        return openpyxl.load_workbook(excel_path)
    return None

def read_edsd_scores(workbook):
    if workbook is None:
        return {}
    sheet = workbook.active
    scores = {}
    
    # 3행의 B열 값으로 어떤 열을 읽을지 결정
    header_b = sheet.cell(row=3, column=2).value
    header_c = sheet.cell(row=3, column=3).value
    
    print(f"DEBUG: header_b = '{header_b}', header_c = '{header_c}'")
    print(f"DEBUG: header_b == '병실' 결과: {header_b == '병실'}")
    
    if header_b == "병실":
        # 기존 형식: B열=병실, C열=점수
        min_col, max_col = 2, 3
        print(f"기존 형식 사용: B열({header_b}), C열({header_c})")
    else:
        # 새로운 형식: C열=병실, D열=점수
        min_col, max_col = 3, 4
        print(f"새로운 형식 사용: C열({header_c}), D열({sheet.cell(row=3, column=4).value})")
    
    for row in sheet.iter_rows(min_row=4, min_col=min_col, max_col=max_col, values_only=True):
        room, score = row
        if room is None or score is None:
            continue
            
        # 병실 형식이 "211_1" 또는 "211_1_jj" 패턴이 아니면 건너뛰기
        if not re.match(r'^\d+_\d+(?:_[a-z]+)?$', str(room)):
            continue
            
        match = re.match(r'(\d+_\d+)(?:_(\w+))?', room)
        if match:
            room_num = match.group(1).replace('_', '-')
            hospital_code = match.group(2) if match.group(2) else "yn"
            ward = room_num.split('-')[0]
            scores.setdefault(hospital_code, {}).setdefault(ward, {})[room_num] = score
    
    return scores

# --- 30일간의 엑셀 파일 찾기 ---
def find_excel_files_in_date_range(base_dir, days_back=30):
    excel_files = []
    today = datetime.now().date()
    for i in range(days_back, -1, -1):
        target_date = today - timedelta(days=i)
        folder_name = target_date.strftime('%Y_%m_%d')
        excel_path = os.path.join(base_dir, 'excels', folder_name, 'edsd_score_output.xlsx')
        if os.path.exists(excel_path):
            excel_files.append((target_date, excel_path))
            print(f"발견된 파일: {folder_name}/edsd_score_output.xlsx")
    return excel_files

# --- 누적 데이터 저장/불러오기 ---
def load_data_store(file_path="data_store.pkl"):
    if os.path.exists(file_path):
        with open(file_path, "rb") as f:
            data = pickle.load(f)
            if 'check_delete_date_flag' not in data:
                data['check_delete_date_flag'] = 0
            return data
    return {'check_delete_date_flag': 0}

def save_data_store(data_store, file_path="data_store.pkl"):
    with open(file_path, "wb") as f:
        pickle.dump(data_store, f)

def delete_data_store(file_path="data_store.pkl"):
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"data_store.pkl 파일이 삭제되었습니다.")

# --- 누적 데이터 오래된 것 삭제 ---
def prune_old_data(accumulated_data, days_back=30):
    today = datetime.now().date()
    cutoff = today - timedelta(days=days_back)
    for hospital_code, wards in accumulated_data.items():
        for ward, rooms in wards.items():
            for room_num, xy_data in rooms.items():
                accumulated_data[hospital_code][ward][room_num] = [
                    (d, s) for d, s in xy_data if d >= cutoff
                ]
    #print(f"{days_back}일 이전 데이터 삭제 완료")
    return accumulated_data

# --- 그래프 생성 (오늘 기준 days_back만큼만) ---
def create_plots_for_date(accumulated_data, current_date, days_back, save_dir="plotImg", axis_dates=None): # days_back은 오늘 날짜 기준으로 몇일을 그래프 그릴지 세팅 값
    os.makedirs(save_dir, exist_ok=True)
    image_files = {}
    colors = ['r', 'b', 'g', 'orange', 'purple', 'pink']

    today = current_date
    start_date = today - timedelta(days=days_back)

    for hospital_code, wards in accumulated_data.items():
        image_files.setdefault(hospital_code, {})
        for ward, rooms in wards.items():
            if not isinstance(rooms, dict):
                continue

            ward_shown = set()
            fig, ax = plt.subplots(figsize=(20, 5))
            ax.set_title(f"{hospital_code} - Room{ward} (~ {current_date.strftime('%Y-%m-%d')})")
            ax.set_xlabel("")
            ax.set_ylabel("EDSD SCORE")

            all_y_values = []

            # 1) 먼저 병실별로 필터링된 데이터와 전체 날짜 집합을 수집
            room_to_filtered = {}
            ward_date_set = set()
            for room_num, xy_data in rooms.items():
                xy_data_filtered = [(d, s) for d, s in xy_data if start_date <= d <= today]
                if xy_data_filtered:
                    room_to_filtered[room_num] = xy_data_filtered
                    for d, _ in xy_data_filtered:
                        ward_date_set.add(d)

            # 날짜를 압축 축으로 사용: 인덱스 0..N-1에 실제 날짜를 매핑
            # 축 라벨은 기본적으로 axis_dates(폴더 존재 날짜)를 사용하고, 없으면 병실 데이터의 날짜를 사용
            if axis_dates is not None:
                axis_dates_filtered = sorted(d for d in axis_dates if start_date <= d <= today)
            else:
                axis_dates_filtered = sorted(ward_date_set)
            date_to_index = {d: i for i, d in enumerate(axis_dates_filtered)}

            # 2) 인덱스 기반으로 선/마커를 그림 (빈 날짜는 축에서 제거되어 압축됨)
            for idx, (room_num, xy_data_filtered) in enumerate(room_to_filtered.items()):
                x_idx = [date_to_index[d] for d, s in xy_data_filtered]
                y = [s + 0.1 if s == 0 else s for d, s in xy_data_filtered]
                all_y_values.extend(y)

                ax.plot(x_idx, y, marker='o', label=room_num, color=colors[idx % len(colors)])

               # 점 위에 값 표시 (인덱스 좌표 기준)
                for xi, yi, yi_orig in zip(x_idx, y, [s for d, s in xy_data_filtered]):
                    key = (xi, yi_orig, ward)
                    if key not in ward_shown:
                        try:
                            x_pos = float(xi) + 0.05
                            y_pos = float(yi) + 0.15
                        except (ValueError, TypeError):
                            x_pos = xi
                            y_pos = yi + 0.15

                        ax.text(x_pos, y_pos, f"{yi_orig}", fontsize=7,
                                ha='left', va='bottom', rotation=0, color='black')
                        ward_shown.add(key)

            if all_y_values:
                ax.set_ylim(0, max(all_y_values) + 3)

            # 3) x축: 인덱스 눈금과 날짜 문자열 라벨 적용, 좌우 여백 0.5 유지
            if axis_dates is not None:
                label_dates = axis_dates_filtered
            else:
                label_dates = sorted(ward_date_set)
            if label_dates:
                ax.set_xticks(list(range(len(label_dates))))
                ax.set_xticklabels([d.strftime('%Y-%m-%d') for d in label_dates], rotation=45, ha='right')
                ax.set_xlim(-0.5, len(label_dates) - 0.5)

            ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))

            img_path = os.path.join(save_dir, f"{hospital_code}_{ward}_{current_date.strftime('%Y%m%d')}.png")
            fig.savefig(img_path, bbox_inches='tight')
            plt.close(fig)

            image_files[hospital_code][ward] = img_path

    return image_files

# --- 엑셀에 그래프 삽입 ---
def save_to_excel(image_files, excel_path):
    """기존 엑셀 파일에 그래프 추가"""
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        print(f"엑셀 파일이 존재하지 않음: {excel_path}")
        return

    # hospital_code -> 시트 이름 매핑
    sheet_name_map = {
        'yn': "영남(경산)",
        'jj': "전남제일(화순)",
        'h': "효사랑(영천)",
        'gj': "구미제일(구미)"
    }

    for hospital_code, wards in image_files.items():
        # hospital_code에 맞는 시트 이름
        sheet_name = sheet_name_map.get(hospital_code, hospital_code)

        # 기존 시트가 있으면 그대로 사용, 없으면 새로 생성
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 기존 이미지들 삭제 (그래프 업데이트를 위해)
            ws._images.clear()
        else:
            ws = wb.create_sheet(title=sheet_name)

        # 이미지 삽입
        row_pos = 1  # 맨 위부터 삽입
        for ward, img_path in wards.items():
            img = ExcelImage(img_path)
            ws.add_image(img, f"A{row_pos}")
            row_pos += 24  # 이미지 간격

    wb.save(excel_path)
    print(f"엑셀 업데이트 완료: {excel_path}")


# --- 메인 실행 ---
if __name__ == '__main__':
    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_files = find_excel_files_in_date_range(base_dir, days_back=30)

    if not excel_files:
        print("30일 내에 존재하는 엑셀 파일이 없습니다.")
        exit()

    data_store = load_data_store()
    accumulated_data = {}

    print(f"\n각 날짜별 그래프 생성 시작... (총 {len(excel_files)}개 파일)")

    # 폴더가 존재하는 날짜들(축 라벨용)
    axis_dates = [d for d, _ in excel_files]

    for i, (current_date, excel_path) in enumerate(excel_files):
        print(f"\n[{i+1}/{len(excel_files)}] 처리중: {current_date} - {os.path.basename(os.path.dirname(excel_path))}")

        wb = open_edsd_score_excel(excel_path)
        current_scores = read_edsd_scores(wb)
        if wb:
            wb.close()

        for hospital_code, wards in current_scores.items():
            accumulated_data.setdefault(hospital_code, {})
            for ward, rooms in wards.items():
                accumulated_data[hospital_code].setdefault(ward, {})
                for room_num, score in rooms.items():
                    accumulated_data[hospital_code][ward].setdefault(room_num, [])
                    accumulated_data[hospital_code][ward][room_num].append((current_date, score))

        accumulated_data = prune_old_data(accumulated_data, days_back=30)

        # 오늘 기준 days_back만큼만 그래프 생성
        save_dir = f"plotImg_{current_date.strftime('%Y%m%d')}"
        image_files = create_plots_for_date(accumulated_data, current_date, days_back=30, save_dir=save_dir, axis_dates=axis_dates)# days_back은 오늘 날짜 기준으로 몇일을 그래프 그릴지 세팅 값

        if image_files:
            save_to_excel(image_files, excel_path)

        import shutil
        if os.path.exists(save_dir):
            shutil.rmtree(save_dir)

    data_store['check_delete_date_flag'] += 1
    #print(f"\n실행 횟수: {data_store['check_delete_date_flag']}/30")

    for hospital_code, wards in accumulated_data.items():
        data_store.setdefault(hospital_code, {})
        for ward, rooms in wards.items():
            data_store[hospital_code].setdefault(ward, {})
            for room_num, xy_data in rooms.items():
                data_store[hospital_code][ward][room_num] = xy_data

    save_data_store(data_store)
    print("\n모든 날짜별 그래프 생성 및 저장 완료!")
